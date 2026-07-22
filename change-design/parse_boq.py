"""
變更設計工具 - 第一步：把台水公司「詳細價目表」原始分頁格式，
轉換成連續的階層完整編碼表（<壹><一><A><1>[1]a1-1 這種格式），
並在每一層自動插入小計。
"""
import re
import openpyxl
from openpyxl.utils import get_column_letter

BIG_CN = '壹貳參叁肆伍陸柒捌玖拾'
SMALL_CN = '一二三四五六七八九十'

FIXED_TIER = {'CB': 1, 'CS': 2, 'UL': 3, 'DB': 4, 'SB': 5}


def classify(a):
    """回傳 (kind, symbol_repr) 或 None(表示不是一個標準代號格式)"""
    if a is None:
        return None
    a = str(a).strip()
    if not a:
        return None
    m = re.fullmatch(r'<([^<>])>', a)
    if m:
        ch = m.group(1)
        if ch in BIG_CN:
            return ('CB', a)
        if ch in SMALL_CN:
            return ('CS', a)
        if re.fullmatch(r'[A-Z]', ch):
            return ('UL', a)
        if re.fullmatch(r'\d+', ch):
            return ('DB', a)
        return ('UNK', a)
    m = re.fullmatch(r'\[(\d+)\]', a)
    if m:
        return ('SB', a)
    m = re.fullmatch(r'[a-z]', a)
    if m:
        return ('BL', a)
    m = re.fullmatch(r'\d+', a)
    if m:
        return ('BD', a)
    return ('UNK', a)


def is_boilerplate(a, b, c, d, e, f, g):
    if all(v is None for v in (a, b, c, d, e, f, g)):
        return True
    if a == '台灣自來水公司':
        return True
    if a == '詳細價目表[契約]':
        return True
    if a == '工程名稱':
        return True
    if a == '施工地點':
        return True
    if a == '項 次':
        return True
    if isinstance(a, str) and '投標廠商' in a:
        return True
    if isinstance(b, str) and '投標廠商' in b:
        return True
    if isinstance(f, str) and '頁' in f and '共' in f:
        return True
    return False


def is_discard_rollup(a, c, d, e, f):
    """原始檔案裡既有的小計/合計列 -> 我們會自己重新計算，所以丟棄。"""
    return a is None and c is None and d is None and e is None and f is not None


class Node:
    __slots__ = ('kind', 'code', 'desc', 'tier', 'is_leaf', 'unit', 'qty',
                 'price', 'total', 'remark', 'children', 'last_child_kind',
                 'last_child_val', 'sum', 'row_no')

    def __init__(self, kind, code, desc, tier, is_leaf, unit=None, qty=None,
                 price=None, total=None, remark=None):
        self.kind = kind
        self.code = code
        self.desc = desc
        self.tier = tier
        self.is_leaf = is_leaf
        self.unit = unit
        self.qty = qty
        self.price = price
        self.total = total
        self.remark = remark
        self.children = []  # list of Node (only for headers)
        self.last_child_kind = None
        self.last_child_val = None
        self.sum = 0
        self.row_no = None  # 之後輸出時填入


def seq_ord(kind, raw):
    if kind == 'BD':
        return int(raw)
    if kind == 'BL':
        return ord(raw) - ord('a') + 1
    return None


def parse_workbook(path, sheet_name='契約詳細表'):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    root_children = []
    stack = []  # list of Node currently open
    validation_log = []
    unnumbered_log = []
    unknown_log = []

    def close_node(node):
        # 把自己的 sum 加到父層（若有）
        if stack:
            parent = stack[-1]
            parent.sum += node.sum
            parent.children.append(node)
        else:
            root_children.append(node)

    def attach_leaf(parent, leaf):
        if parent is not None:
            parent.sum += leaf.total if leaf.total else 0
            parent.children.append(leaf)
        else:
            root_children.append(leaf)

    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value

        if is_boilerplate(a, b, c, d, e, f, g):
            continue

        if is_discard_rollup(a, c, d, e, f):
            # 驗證：目前最上層(尚未關閉)的累積總和應該要等於原始檔案自己算的小計/計
            expected = stack[-1].sum if stack else None
            validation_log.append({
                'row': r, 'label': b, 'raw_value': f, 'computed_so_far': expected
            })
            continue

        cls = classify(a)
        # 工項判準：F(複價)有值，或 有單位C且有數量D（小計列已於上面排除）
        is_leaf_row = (f is not None) or (c is not None and d is not None)

        if a is None:
            # ---- 未編號列 ----
            unnumbered_log.append({'row': r, 'a': a, 'b': b, 'c': c, 'd': d,
                                    'e': e, 'f': f, 'is_header': not is_leaf_row})
            if is_leaf_row:
                parent = stack[-1] if stack else None
                leaf = Node('UN', (parent.code if parent else '') , b, 
                            (parent.tier + 1) if parent else 1, True,
                            unit=c, qty=d, price=e, total=f, remark=g)
                leaf.row_no = r
                attach_leaf(parent, leaf)
            else:
                # 未編號標題：視為與目前開啟中的未編號標題平行(攤平一層)
                while stack and stack[-1].kind == 'UN':
                    close_node(stack.pop())
                parent = stack[-1] if stack else None
                node = Node('UN', (parent.code if parent else '') + f'《{b}》',
                            b, (parent.tier + 1) if parent else 1, False)
                stack.append(node)
            continue

        kind, raw = cls

        if kind == 'UNK':
            unknown_log.append({'row': r, 'a': a, 'b': b})
            # 當作未編號列處理，附掛在目前最上層
            parent = stack[-1] if stack else None
            if is_leaf_row:
                leaf = Node('UNK', (parent.code if parent else '') + str(a), b,
                            (parent.tier + 1) if parent else 1, True,
                            unit=c, qty=d, price=e, total=f, remark=g)
                leaf.row_no = r
                attach_leaf(parent, leaf)
            else:
                node = Node('UNK', (parent.code if parent else '') + str(a), b,
                            (parent.tier + 1) if parent else 1, False)
                stack.append(node)
            continue

        if kind in FIXED_TIER:
            tier = FIXED_TIER[kind]
            while stack and stack[-1].tier >= tier:
                close_node(stack.pop())
            parent = stack[-1] if stack else None
            if kind in ('CB', 'CS', 'UL'):
                code = raw
            else:  # DB
                code = (parent.code if parent else '') + raw
            if is_leaf_row:
                leaf = Node(kind, code, b, tier, True, unit=c, qty=d, price=e,
                            total=f, remark=g)
                leaf.row_no = r
                attach_leaf(parent, leaf)
            else:
                node = Node(kind, code, b, tier, False)
                stack.append(node)
            continue

        # ---- SEQ zone: BL / BD ----
        val = seq_ord(kind, raw)
        while True:
            if not stack:
                # 沒有任何父層，當作根節點的子項
                parent = None
                code = raw
                tier = 1
                break
            top = stack[-1]
            if top.last_child_kind is None:
                top.last_child_kind = kind
                top.last_child_val = val
                parent = top
                break
            elif top.last_child_kind == kind and val == top.last_child_val + 1:
                top.last_child_val = val
                parent = top
                break
            else:
                close_node(stack.pop())
                continue

        if parent is not None:
            tier = parent.tier + 1
            if kind == 'BD' and parent.kind == 'BD':
                code = parent.code + '-' + raw
            else:
                code = parent.code + raw

        if is_leaf_row:
            leaf = Node(kind, code, b, tier, True, unit=c, qty=d, price=e,
                        total=f, remark=g)
            leaf.row_no = r
            attach_leaf(parent, leaf)
        else:
            node = Node(kind, code, b, tier, False)
            stack.append(node)

    # 收尾：把 stack 剩下的節點全部關閉
    while stack:
        close_node(stack.pop())

    return root_children, validation_log, unnumbered_log, unknown_log


def debug_trace(path, row_start, row_end, sheet_name='契約詳細表'):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    root_children = []
    stack = []

    def close_node(node):
        if stack:
            parent = stack[-1]
            parent.sum += node.sum
            parent.children.append(node)
        else:
            root_children.append(node)
        print(f'    [CLOSE] code={node.code!r} desc={node.desc!r} sum={node.sum}')

    def attach_leaf(parent, leaf):
        if parent is not None:
            parent.sum += leaf.total if leaf.total else 0
            parent.children.append(leaf)
        else:
            root_children.append(leaf)

    for r in range(1, row_end + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value

        if is_boilerplate(a, b, c, d, e, f, g):
            continue
        if is_discard_rollup(a, c, d, e, f):
            if r >= row_start:
                print(f'row={r} [discard 小計] label={b} raw={f} | stack_top_sum={stack[-1].sum if stack else None}')
            continue

        cls = classify(a)
        is_leaf_row = f is not None

        if r >= row_start:
            print(f'row={r} a={a!r} b={b!r} f={f!r} | stack_before=' +
                  str([(n.code, n.kind, n.last_child_kind, n.last_child_val, round(n.sum,2)) for n in stack]))

        if a is None:
            if is_leaf_row:
                parent = stack[-1] if stack else None
                leaf = Node('UN', '', b, 0, True, unit=c, qty=d, price=e, total=f, remark=g)
                attach_leaf(parent, leaf)
            else:
                while stack and stack[-1].kind == 'UN':
                    close_node(stack.pop())
                parent = stack[-1] if stack else None
                node = Node('UN', '', b, (parent.tier + 1) if parent else 1, False)
                stack.append(node)
            continue

        kind, raw = cls
        if kind == 'UNK':
            parent = stack[-1] if stack else None
            if is_leaf_row:
                leaf = Node('UNK', str(a), b, 0, True, unit=c, qty=d, price=e, total=f, remark=g)
                attach_leaf(parent, leaf)
            else:
                node = Node('UNK', str(a), b, (parent.tier + 1) if parent else 1, False)
                stack.append(node)
            continue

        if kind in FIXED_TIER:
            tier = FIXED_TIER[kind]
            while stack and stack[-1].tier >= tier:
                close_node(stack.pop())
            parent = stack[-1] if stack else None
            code = raw if kind in ('CB', 'CS', 'UL') else (parent.code if parent else '') + raw
            if is_leaf_row:
                leaf = Node(kind, code, b, tier, True, unit=c, qty=d, price=e, total=f, remark=g)
                attach_leaf(parent, leaf)
            else:
                node = Node(kind, code, b, tier, False)
                stack.append(node)
            continue

        val = seq_ord(kind, raw)
        while True:
            if not stack:
                parent = None
                code = raw
                tier = 1
                break
            top = stack[-1]
            if top.last_child_kind is None:
                top.last_child_kind = kind
                top.last_child_val = val
                parent = top
                break
            elif top.last_child_kind == kind and val == top.last_child_val + 1:
                top.last_child_val = val
                parent = top
                break
            else:
                close_node(stack.pop())
                continue
        if parent is not None:
            tier = parent.tier + 1
            code = parent.code + ('-' if (kind == 'BD' and parent.kind == 'BD') else '') + raw
        if is_leaf_row:
            leaf = Node(kind, code, b, tier, True, unit=c, qty=d, price=e, total=f, remark=g)
            attach_leaf(parent, leaf)
        else:
            node = Node(kind, code, b, tier, False)
            stack.append(node)


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else '/home/claude/詳細表.xlsx'
    roots, vlog, unlog, unklog = parse_workbook(path)

    print(f'共 {len(roots)} 個頂層節點')
    print(f'驗證用的原始小計/計 列，共 {len(vlog)} 筆')
    mismatches = 0
    for v in vlog:
        exp = v['computed_so_far']
        raw = v['raw_value']
        ok = exp is not None and abs(exp - raw) < 0.05
        if not ok:
            mismatches += 1
            print(f"  [不符] row={v['row']} label={v['label']!r} raw={raw} computed={exp}")
    print(f'驗證結果：{len(vlog)-mismatches}/{len(vlog)} 相符')

    print(f'\n未編號列數量: {len(unlog)}')
    for u in unlog[:20]:
        print(' ', u)
    if len(unlog) > 20:
        print(f'  ...(共 {len(unlog)} 筆，只列出前20筆)')

    print(f'\n未知格式列數量: {len(unklog)}')
    for u in unklog[:20]:
        print(' ', u)
