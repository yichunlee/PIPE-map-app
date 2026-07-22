"""
變更設計核心邏輯（不含介面，可獨立測試/呼叫）

概念：
- 從原契約詳細表解析出階層樹（沿用 parse_boq）。
- 「群組」= 底下直接掛有明細項目(leaf)的標題節點，例如 <A><1>[1] 整地及拆除工程。
  變更設計明細表就是以群組為單位，展開成 一/二/三/四 四種狀態 + 合計。
- 使用者的變更以兩種資料表示：
    changes    : {項次code -> 變更後數量}   （原契約項目改數量）
    new_items  : [ NewItem, ... ]            （新增項目，含單價分析）

輸出：完整變更設計明細表 .xlsx（欄位比照範例）。
"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from parse_boq import parse_workbook, is_boilerplate, is_discard_rollup

import re as _re

_CN_BIG = '壹貳參叁肆伍陸柒捌玖拾'
_CN_SM = '一二三四五六七八九十'


def code_sort_key(code):
    """把項次完整編碼轉成可排序的序列（同層兄弟依編碼大小排序）。"""
    if not code:
        return []
    tokens = _re.findall(r'<[^<>]+>|\[\d+\]|[a-z]|\d+|《[^》]*》', code)
    key = []
    for t in tokens:
        if t.startswith('<') and t.endswith('>'):
            ch = t[1:-1]
            if ch in _CN_BIG:
                key.append((0, _CN_BIG.index(ch)))
            elif ch in _CN_SM:
                key.append((1, _CN_SM.index(ch)))
            elif _re.fullmatch(r'[A-Z]+', ch):
                key.append((2, ch))
            elif ch.isdigit():
                key.append((3, int(ch)))
            else:
                key.append((9, ch))
        elif t.startswith('['):
            key.append((4, int(t[1:-1])))
        elif _re.fullmatch(r'[a-z]', t):
            key.append((5, t))
        elif t.isdigit():
            key.append((6, int(t)))
        else:
            key.append((9, t))
    return key


# ----------------------------------------------------------------------------
# 資料模型
# ----------------------------------------------------------------------------
class Leaf:
    """一個原契約明細項目（唯讀基準）。"""
    __slots__ = ('code', 'desc', 'unit', 'orig_qty', 'price', 'orig_total',
                 'remark', 'group_code', 'row_no')

    def __init__(self, code, desc, unit, qty, price, total, remark, group_code, row_no=None):
        self.code = code
        self.desc = desc
        self.unit = unit
        self.orig_qty = qty if qty is not None else 0
        self.price = price
        self.orig_total = total if total is not None else 0
        self.remark = remark
        self.group_code = group_code
        self.row_no = row_no


class Group:
    """一個變更設計的分組單位（例如 整地及拆除工程）。"""
    __slots__ = ('code', 'desc', 'leaves', 'ancestors')

    def __init__(self, code, desc, ancestors):
        self.code = code
        self.desc = desc
        self.leaves = []          # list[Leaf]
        self.ancestors = ancestors  # list[(code, desc)] 由外到內的祖先標題


def _clean(s):
    """移除字串換行，避免匯出時儲存格變多行、表格列被截斷。"""
    if s is None:
        return ''
    return str(s).replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ').strip()


# 物價指數調整(物調)適用的材料類別 — 供單價分析「備註」下拉選單
MATERIAL_TYPES = ['預拌混凝土', '鋼筋', '鋼板', '型鋼', '瀝青混凝土',
                  '水泥及其製品類', '金屬製品類', '瀝青及其製品類']


class UPARow:
    """單價分析的一列工料。"""
    __slots__ = ('desc', 'unit', 'qty', 'price', 'remark')

    def __init__(self, desc='', unit='', qty=0.0, price=0.0, remark=''):
        self.desc = _clean(desc)
        self.unit = _clean(unit)
        self.qty = qty
        self.price = price
        self.remark = _clean(remark)   # 材料類別(物調)，空字串=不適用

    @property
    def total(self):
        return round((self.qty or 0) * (self.price or 0), 2)

    def to_dict(self):
        return {'desc': self.desc, 'unit': self.unit, 'qty': self.qty,
                'price': self.price, 'remark': self.remark}


class NewItem:
    """一個新增項目（含單價分析：原契約部份 + 新增部份）。"""
    def __init__(self, group_code, code, desc, unit, qty):
        self.group_code = group_code
        self.code = _clean(code)
        self.desc = _clean(desc)
        self.unit = _clean(unit)
        self.qty = qty
        self.upa_orig = []   # list[UPARow]  原契約單價部份
        self.upa_new = []    # list[UPARow]  新增單價部份
        self.reason = ''     # 新增項目原因分析（匯出到分析表三）

    @property
    def orig_unit_price(self):
        return round(sum(r.total for r in self.upa_orig), 2)

    @property
    def new_unit_price(self):
        return round(sum(r.total for r in self.upa_new), 2)

    def to_dict(self):
        return {
            'group_code': self.group_code, 'code': self.code, 'desc': self.desc,
            'unit': self.unit, 'qty': self.qty, 'reason': self.reason,
            'upa_orig': [r.to_dict() for r in self.upa_orig],
            'upa_new': [r.to_dict() for r in self.upa_new],
        }

    @staticmethod
    def from_dict(d):
        it = NewItem(d['group_code'], d['code'], d['desc'], d['unit'], d['qty'])
        it.reason = _clean(d.get('reason', ''))
        it.upa_orig = [UPARow(**r) for r in d.get('upa_orig', [])]
        it.upa_new = [UPARow(**r) for r in d.get('upa_new', [])]
        return it


# ----------------------------------------------------------------------------
# 從原契約樹建立 群組 / 明細
# ----------------------------------------------------------------------------
def build_groups(roots):
    """回傳 (groups_in_order, leaf_by_code, group_by_code)。"""
    groups = []
    leaf_by_code = {}
    group_by_code = {}

    def make_group(code, desc, ancestors, leaves_src):
        g = Group(code, desc, list(ancestors))
        for lf in leaves_src:
            leaf = Leaf(lf.code, lf.desc, lf.unit, lf.qty, lf.price,
                        lf.total, lf.remark, code, getattr(lf, 'row_no', None))
            g.leaves.append(leaf)
            leaf_by_code[lf.code] = leaf
        groups.append(g)
        group_by_code[code] = g
        return g

    def walk(node, ancestors):
        if node.is_leaf:
            return
        # 依編碼把子節點排序（同層兄弟：<A> <B> <C> <D> <E> 依序）
        node.children.sort(key=lambda c: code_sort_key(c.code))
        direct_leaves = [c for c in node.children if c.is_leaf]
        direct_headers = [c for c in node.children if not c.is_leaf]

        if direct_leaves and not direct_headers:
            # 純明細容器：整個節點成一組（例如 整地及拆除工程）
            make_group(node.code, node.desc, ancestors, node.children)
        elif direct_leaves:
            # 混合節點（同時有直接明細與子標題，例如 <B>職業安全衛生費 或 h2）：
            # 每個直接明細各自成一個「單項群組」(用自己的葉編碼與名稱)，
            # 子標題再各自遞迴。這樣 <B><2>其他安全設施及作業費 會像 <D><E> 一樣
            # 以獨立列顯示(含合計)，且與子標題共用同一個父層小計。
            child_anc = ancestors + [(node.code, node.desc)]
            for c in node.children:
                if c.is_leaf:
                    make_group(c.code, c.desc, child_anc, [c])
                else:
                    walk(c, child_anc)
        else:
            # 純標題：逐一遞迴
            child_anc = ancestors + [(node.code, node.desc)]
            for c in node.children:
                walk(c, child_anc)

    for r in sorted(roots, key=lambda n: code_sort_key(n.code)):
        if r.is_leaf:
            # 頂層就是明細（例如 <柒>營業稅）：自成一個單項群組（無祖先）
            g = Group(r.code, r.desc, [])
            leaf = Leaf(r.code, r.desc, r.unit, r.qty, r.price, r.total,
                        r.remark, r.code, getattr(r, 'row_no', None))
            g.leaves.append(leaf)
            groups.append(g)
            group_by_code[r.code] = g
            leaf_by_code[r.code] = leaf
        else:
            walk(r, [(r.code, r.desc)])
    return groups, leaf_by_code, group_by_code


class ChangeModel:
    def __init__(self, src_path):
        self.src_path = src_path      # 保留原始檔路徑，供產生「變更後詳細價目表」時重用其 A 欄編碼
        roots, _, _, _ = parse_workbook(src_path)
        self.roots = roots
        self.groups, self.leaf_by_code, self.group_by_code = build_groups(roots)
        self.changes = {}      # code -> new_qty
        self.new_items = []    # list[NewItem]
        self.rate_amounts = {} # code -> {'inc': 增加金額, 'dec': 減少金額}（費率型項目手填）
        self.reasons = {}      # code -> 數量增加/減少原因分析（原契約工項用）

    # ---- 變更操作 ----
    def set_new_qty(self, code, new_qty):
        if code not in self.leaf_by_code:
            raise KeyError(f'找不到項次 {code}')
        self.changes[code] = float(new_qty)

    def clear_change(self, code):
        self.changes.pop(code, None)
        self.rate_amounts.pop(code, None)

    def set_rate_amount(self, code, inc=0.0, dec=0.0):
        """費率型項目(單位有、單價為'--')手動填入增加/減少金額。"""
        inc = float(inc or 0); dec = float(dec or 0)
        if inc == 0 and dec == 0:
            self.rate_amounts.pop(code, None)
        else:
            self.rate_amounts[code] = {'inc': inc, 'dec': dec}

    def add_new_item(self, item: NewItem):
        self.new_items.append(item)

    # ---- 儲存 / 讀取變更狀態 ----
    def save_state(self, path):
        data = {
            'changes': self.changes,
            'new_items': [it.to_dict() for it in self.new_items],
            'rate_amounts': self.rate_amounts,
            'reasons': self.reasons,
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def load_state(self, path):
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        self.changes = {k: float(v) for k, v in data.get('changes', {}).items()}
        self.new_items = [NewItem.from_dict(d) for d in data.get('new_items', [])]
        self.rate_amounts = {k: {'inc': float(v.get('inc', 0)), 'dec': float(v.get('dec', 0))}
                             for k, v in data.get('rate_amounts', {}).items()}
        self.reasons = {k: str(v) for k, v in data.get('reasons', {}).items() if str(v).strip()}


def is_rate_item(lf):
    """費率型項目：有單位但單價不是數字(通常為 '--')，金額放在 orig_total。"""
    return not isinstance(lf.price, (int, float))


# ----------------------------------------------------------------------------
# 產生變更設計明細表
# ----------------------------------------------------------------------------
COLS = ['項次', '項目及說明', '單位', '原定數量', '變更後數量', '增減數量',
        '單價', '前次修正預算', '第N次變更設計', '增加金額', '減少金額']
# 這裡「第十二次修正預算 / 第一次變更設計」的文字可由呼叫端覆寫

C_CODE, C_DESC, C_UNIT, C_OQTY, C_NQTY, C_DQTY, C_PRICE, C_BEFORE, C_AFTER, C_INC, C_DEC = range(1, 12)


def _newitem_rel_code(model, it):
    """新增項目在詳細表 A 欄要放的『相對代號』。
    新增項目編號 = 所屬 group 完整編碼 + 尾碼，尾碼即相對代號（如 <A><1>[1]16 → 16）。"""
    gcode = it.group_code or ''
    if it.code.startswith(gcode):
        rel = it.code[len(gcode):]
        return rel.lstrip('-') if rel.startswith('-') else rel
    return _short_code(it.code)


def generate_detail_boq(model: ChangeModel, out_path, title_suffix='（第一次變更設計後）'):
    """產生『變更後詳細價目表』，格式與台水原契約詳細表相同，可被 parse_workbook 再次讀入。

    策略（零風險、無公式快取問題）：
      逐列掃描原始檔，把每一列原樣搬到新檔（A 欄相對代號、B 說明、C 單位直接沿用原檔值），
      唯一改動：工項列的 D(數量) 換成變更後數量、E(單價)/F(複價) 一律寫『具體數值』(不留公式)。
      小計/合計列丟棄（parse 會自行重算）。每個分組的新增項目，插在該組最後一列工項之後。
    """
    import openpyxl as _oxl
    from openpyxl.styles import Alignment as _Al, Font as _Ft
    src = _oxl.load_workbook(model.src_path, data_only=True)   # data_only：讀到公式的快取值
    sws = src['契約詳細表'] if '契約詳細表' in src.sheetnames else src[src.sheetnames[0]]

    # row_no -> leaf（用於取變更後數量、單價）
    leaf_by_row = {lf.row_no: lf for lf in model.leaf_by_code.values() if lf.row_no}
    # group 最後一列 -> 該組新增項目
    news_by_last_row = {}
    for it in model.new_items:
        g = model.group_by_code.get(it.group_code)
        rows = [lf.row_no for lf in (g.leaves if g else []) if lf.row_no]
        if rows:
            news_by_last_row.setdefault(max(rows), []).append(it)

    out = _oxl.Workbook()
    ows = out.active
    ows.title = '契約詳細表'
    left = _Al(horizontal='left', vertical='center', wrap_text=True)
    right = _Al(horizontal='right', vertical='center')
    center = _Al(horizontal='center', vertical='center')

    def newitem_unit_price(it):
        rows = ([d.to_dict() for d in it.upa_orig] + [d.to_dict() for d in it.upa_new]) \
            if hasattr(it, 'upa_orig') else []
        return round(sum((r.get('qty', 0) or 0) * (r.get('price', 0) or 0) for r in rows), 2)

    orow = 0
    for sr in range(1, sws.max_row + 1):
        a = sws.cell(sr, 1).value
        b = sws.cell(sr, 2).value
        c = sws.cell(sr, 3).value
        d = sws.cell(sr, 4).value
        e = sws.cell(sr, 5).value
        f = sws.cell(sr, 6).value
        g = sws.cell(sr, 7).value

        if is_boilerplate(a, b, c, d, e, f, g):
            # 表頭原樣保留（給人看，parse 會跳過）
            orow += 1
            for col, val in enumerate([a, b, c, d, e, f, g], start=1):
                if val is not None:
                    ows.cell(orow, col, val)
            continue
        if is_discard_rollup(a, c, d, e, f):
            continue   # 小計/合計列丟棄

        lf = leaf_by_row.get(sr)
        if lf is not None:
            # 工項列：A/B/C 沿用原檔，D 換變更後數量，E/F 寫具體值
            orow += 1
            qty = model.changes.get(lf.code, lf.orig_qty)
            price = lf.price
            ows.cell(orow, 1, a); ows.cell(orow, 1).alignment = center
            ows.cell(orow, 2, b); ows.cell(orow, 2).alignment = left
            ows.cell(orow, 3, c); ows.cell(orow, 3).alignment = center
            dcell = ows.cell(orow, 4, qty); dcell.alignment = right
            if isinstance(price, (int, float)):
                ows.cell(orow, 5, price).alignment = right
                ows.cell(orow, 6, round(qty * price, 2)).alignment = right
            else:
                # 費率型：單價 '--'，複價沿用原檔快取值（或原契約複價）
                ows.cell(orow, 5, price if price is not None else '--').alignment = center
                ows.cell(orow, 6, f if f is not None else lf.orig_total).alignment = right
            if g is not None:
                ows.cell(orow, 7, g).alignment = left
        else:
            # 階層標題列（A 欄是代號、無 F 值）：原樣搬
            orow += 1
            for col, val in enumerate([a, b, c, d, e, f, g], start=1):
                if val is not None:
                    ows.cell(orow, col, val)
            ows.cell(orow, 1).alignment = center
            ows.cell(orow, 2).alignment = left

        # 這列是某組最後一列工項 → 接著插入該組新增項目
        if sr in news_by_last_row:
            for it in news_by_last_row[sr]:
                orow += 1
                up = newitem_unit_price(it)
                qty = float(it.qty or 0)
                ows.cell(orow, 1, _newitem_rel_code(model, it)).alignment = center
                ows.cell(orow, 2, it.desc).alignment = left
                ows.cell(orow, 3, it.unit).alignment = center
                ows.cell(orow, 4, qty).alignment = right
                ows.cell(orow, 5, up).alignment = right
                ows.cell(orow, 6, round(up * qty, 2)).alignment = right

    # 表頭標注變更後版本
    for rr in range(1, min(orow, 6) + 1):
        v = ows.cell(rr, 1).value
        if v and '詳細價目表' in str(v):
            ows.cell(rr, 1, str(v) + '  ' + title_suffix)
            break

    # 欄寬
    for col, w in zip('ABCDEFG', [16, 46, 6, 11, 12, 14, 12]):
        ows.column_dimensions[col].width = w

    out.save(out_path)
    return out_path


def generate_change_xlsx(model: ChangeModel, out_path,
                         before_label='前次修正預算',
                         after_label='第N次變更設計'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '變更設計明細表'

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(name='Arial', bold=True)
    normal = Font(name='Arial')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    grp_fill = PatternFill('solid', fgColor='FCE4D6')
    sub_fill = PatternFill('solid', fgColor='FFF2CC')
    anc_fill = PatternFill('solid', fgColor='F2F2F2')       # 祖先標題底色
    lvl_fill = PatternFill('solid', fgColor='E2EFDA')       # 層級小計底色

    # ---- 表頭（兩列）----
    ws.merge_cells('A1:A2'); ws['A1'] = '項次'
    ws.merge_cells('B1:B2'); ws['B1'] = '項目及說明'
    ws.merge_cells('C1:C2'); ws['C1'] = '單位'
    ws.merge_cells('D1:D2'); ws['D1'] = '原定\n數量'
    ws.merge_cells('E1:E2'); ws['E1'] = '變更後\n數量'
    ws.merge_cells('F1:F2'); ws['F1'] = '增減\n數量'
    ws.merge_cells('G1:G2'); ws['G1'] = '單價'
    ws.merge_cells('H1:H2'); ws['H1'] = before_label
    ws.merge_cells('I1:I2'); ws['I1'] = after_label
    ws.merge_cells('J1:K1'); ws['J1'] = '增減合價'
    ws['J2'] = '增加金額'
    ws['K2'] = '減少金額'
    for row in (1, 2):
        for col in range(1, 12):
            c = ws.cell(row=row, column=col)
            c.font = bold; c.alignment = center; c.border = border; c.fill = hdr_fill

    r = 3  # 目前寫入列

    def money_fmt(cell):
        cell.number_format = '#,##0.00;(#,##0.00);-'

    def qty_fmt(cell):
        cell.number_format = '#,##0.00;(#,##0.00);-'

    def dedup_ancestors(ancestors):
        """去除相鄰重複的祖先 code，回傳 [(code, desc), ...]"""
        out = []
        for code, desc in ancestors:
            if out and out[-1][0] == code:
                continue
            out.append((code, desc))
        return out

    # 層級堆疊：每一項為 {'code','desc','rows':[要加總的子層列號]}
    level_stack = []
    rendered_new_items = set()   # 已在某群組渲染的新增項目 id（安全網用）
    top_level_rows = []   # 最頂層(如 <壹> <柒>)各自的小計列，供總計用

    def emit_level_header(code, desc):
        nonlocal r
        ws.cell(row=r, column=C_CODE, value=code).font = bold
        ws.cell(row=r, column=C_DESC, value=desc).font = bold
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = border
            ws.cell(row=r, column=col).fill = anc_fill
        r += 1

    def emit_level_subtotal(level):
        """輸出某層級的小計列，回傳列號"""
        nonlocal r
        ws.cell(row=r, column=C_CODE, value=level['code']).font = bold
        ws.cell(row=r, column=C_DESC, value=f"{level['desc']}　小計").font = bold
        rows = level['rows']
        for cc, ch in ((C_BEFORE, 'H'), (C_AFTER, 'I'), (C_INC, 'J'), (C_DEC, 'K')):
            c = ws.cell(row=r, column=cc,
                        value=('=' + '+'.join(f'{ch}{i}' for i in rows)) if rows else 0)
            money_fmt(c); c.font = bold
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = border
            ws.cell(row=r, column=col).fill = lvl_fill
        this = r
        r += 1
        return this

    def close_levels_to(depth):
        """把堆疊收斂到指定深度，沿途輸出小計並往上冒泡"""
        while len(level_stack) > depth:
            closing = level_stack.pop()
            st = emit_level_subtotal(closing)
            if level_stack:
                level_stack[-1]['rows'].append(st)
            else:
                top_level_rows.append(st)

    # 哪些 code 同時是「別的群組的祖先」→ 它本身也是一個層級(既有直接明細，又有子標題)
    level_codes = set()
    for g in model.groups:
        for code, _ in dedup_ancestors(g.ancestors):
            level_codes.add(code)

    for g in model.groups:
        g_anc = dedup_ancestors(g.ancestors)
        # 若此群組的 code 本身也是層級(例如 <B> 職業安全衛生費：既有 <B><2> 直接明細，
        # 又有 <B><1> 子標題)，就把它當成一個層級開啟，讓它的直接明細與子標題共用同一個小計。
        is_level_group = g.code in level_codes
        chain = g_anc + [(g.code, g.desc)] if is_level_group else g_anc

        # 找出目前堆疊與此鏈的共同前綴長度
        common = 0
        while (common < len(level_stack) and common < len(chain)
               and level_stack[common]['code'] == chain[common][0]):
            common += 1

        # 收掉比共同前綴更深的層級（會輸出它們的小計）
        close_levels_to(common)

        # 開啟共同前綴以下的新層級（輸出標題並入堆疊）
        for i in range(common, len(chain)):
            code, desc = chain[i]
            emit_level_header(code, desc)
            level_stack.append({'code': code, 'desc': desc, 'rows': []})

        # 分類這個群組裡的明細
        decreased, increased, unchanged_total = [], [], 0.0
        changed_orig_sum = 0.0
        for lf in g.leaves:
            if lf.code in model.rate_amounts:
                # 費率型項目：手填增加/減少金額
                ra = model.rate_amounts[lf.code]
                changed_orig_sum += lf.orig_total
                if ra.get('inc', 0) > 0:
                    increased.append((lf, None))     # None 代表費率型
                elif ra.get('dec', 0) > 0:
                    decreased.append((lf, None))
                else:
                    changed_orig_sum -= lf.orig_total
            elif lf.code in model.changes:
                nq = model.changes[lf.code]
                changed_orig_sum += lf.orig_total
                if nq < lf.orig_qty:
                    decreased.append((lf, nq))
                elif nq > lf.orig_qty:
                    increased.append((lf, nq))
                # 等於就當未變動（但已被扣掉，補回）
                else:
                    changed_orig_sum -= lf.orig_total
            # 未變動的稍後用殘差計算
        group_orig_total = sum(lf.orig_total for lf in g.leaves)
        unchanged_before = round(group_orig_total - changed_orig_sum, 2)

        news = [it for it in model.new_items if it.group_code == g.code]
        rendered_new_items.update(id(it) for it in news)

        # ---- 群組標題列（若本身就是層級，層級標題已印，不重複）----
        if not is_level_group:
            ws.cell(row=r, column=C_CODE, value=g.code).font = bold
            ws.cell(row=r, column=C_DESC, value=g.desc).font = bold
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
                ws.cell(row=r, column=col).fill = grp_fill
            r += 1

        section_after_rows = []  # 收集各段「變更設計」金額所在列 -> 供合計 SUM
        section_inc_rows = []
        section_dec_rows = []

        # ---- 一 原契約未變動部份 ----
        ws.cell(row=r, column=C_CODE, value='一').font = normal
        ws.cell(row=r, column=C_DESC, value='原契約未變動部份').font = normal
        cb = ws.cell(row=r, column=C_BEFORE, value=unchanged_before); money_fmt(cb); cb.font = normal
        ca = ws.cell(row=r, column=C_AFTER, value=f'=H{r}'); money_fmt(ca); ca.font = normal
        ci = ws.cell(row=r, column=C_INC, value=f'=IF(I{r}>H{r},I{r}-H{r},0)'); money_fmt(ci); ci.font = normal
        cd = ws.cell(row=r, column=C_DEC, value=f'=IF(H{r}>I{r},H{r}-I{r},0)'); money_fmt(cd); cd.font = normal
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = border
        section_after_rows.append(r); section_inc_rows.append(r); section_dec_rows.append(r)
        r += 1

        def emit_item_line(lf, nq):
            nonlocal r
            ws.cell(row=r, column=C_CODE, value=lf.code).font = normal
            ws.cell(row=r, column=C_DESC, value=lf.desc).font = normal
            ws.cell(row=r, column=C_UNIT, value=lf.unit).font = normal
            if nq is None:   # 費率型項目：單價為 '--'，金額用 orig_total + 手填增減
                ra = model.rate_amounts.get(lf.code, {'inc': 0, 'dec': 0})
                oq = ws.cell(row=r, column=C_OQTY, value=lf.orig_qty); qty_fmt(oq); oq.font = normal
                nqc = ws.cell(row=r, column=C_NQTY, value=lf.orig_qty); qty_fmt(nqc); nqc.font = normal
                ws.cell(row=r, column=C_DQTY, value='').font = normal
                ws.cell(row=r, column=C_PRICE, value='--').font = normal
                hb = ws.cell(row=r, column=C_BEFORE, value=lf.orig_total); money_fmt(hb); hb.font = normal
                ha = ws.cell(row=r, column=C_AFTER, value=f'=H{r}+J{r}-K{r}'); money_fmt(ha); ha.font = normal
                hi = ws.cell(row=r, column=C_INC, value=ra.get('inc', 0)); money_fmt(hi); hi.font = normal
                hd = ws.cell(row=r, column=C_DEC, value=ra.get('dec', 0)); money_fmt(hd); hd.font = normal
            else:
                oq = ws.cell(row=r, column=C_OQTY, value=lf.orig_qty); qty_fmt(oq); oq.font = normal
                nqc = ws.cell(row=r, column=C_NQTY, value=nq); qty_fmt(nqc); nqc.font = normal
                dq = ws.cell(row=r, column=C_DQTY, value=f'=E{r}-D{r}'); qty_fmt(dq); dq.font = normal
                pr = ws.cell(row=r, column=C_PRICE, value=lf.price); pr.number_format = '#,##0.00'; pr.font = normal
                hb = ws.cell(row=r, column=C_BEFORE, value=f'=D{r}*G{r}'); money_fmt(hb); hb.font = normal
                ha = ws.cell(row=r, column=C_AFTER, value=f'=H{r}+J{r}-K{r}'); money_fmt(ha); ha.font = normal
                # 增加金額 = (新-舊) × 單價，但只在增加時 > 0
                hi = ws.cell(row=r, column=C_INC, value=f'=IF(E{r}>D{r},(E{r}-D{r})*G{r},0)'); money_fmt(hi); hi.font = normal
                # 減少金額 = (舊-新) × 單價，但只在減少時 > 0
                hd = ws.cell(row=r, column=C_DEC, value=f'=IF(D{r}>E{r},(D{r}-E{r})*G{r},0)'); money_fmt(hd); hd.font = normal
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
            this = r
            r += 1
            return this

        def emit_subtotal(rows):
            """rows: list of data-row indices that make up this section."""
            nonlocal r
            ws.cell(row=r, column=C_DESC, value='小計').font = bold
            if rows:
                rng = lambda col: '+'.join(f'{col}{i}' for i in rows)
                hb = ws.cell(row=r, column=C_BEFORE, value=f'={rng("H")}')
                ha = ws.cell(row=r, column=C_AFTER, value=f'={rng("I")}')
                hi = ws.cell(row=r, column=C_INC, value=f'={rng("J")}')
                hd = ws.cell(row=r, column=C_DEC, value=f'={rng("K")}')
            else:
                hb = ws.cell(row=r, column=C_BEFORE, value=0)
                ha = ws.cell(row=r, column=C_AFTER, value=0)
                hi = ws.cell(row=r, column=C_INC, value=0)
                hd = ws.cell(row=r, column=C_DEC, value=0)
            for cc in (hb, ha, hi, hd):
                money_fmt(cc); cc.font = bold
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
                ws.cell(row=r, column=col).fill = sub_fill
            this = r
            r += 1
            return this

        # ---- 二 原契約減少部份 ----
        if decreased:
            ws.cell(row=r, column=C_CODE, value='二').font = normal
            ws.cell(row=r, column=C_DESC, value='原契約減少部份').font = normal
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
            r += 1
            drows = [emit_item_line(lf, nq) for lf, nq in decreased]
            st = emit_subtotal(drows)
            section_after_rows.append(st); section_inc_rows.append(st); section_dec_rows.append(st)

        # ---- 三 原契約增加部份 ----
        if increased:
            ws.cell(row=r, column=C_CODE, value='三').font = normal
            ws.cell(row=r, column=C_DESC, value='原契約增加部份').font = normal
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
            r += 1
            irows = [emit_item_line(lf, nq) for lf, nq in increased]
            st = emit_subtotal(irows)
            section_after_rows.append(st); section_inc_rows.append(st); section_dec_rows.append(st)

        # ---- 四 新增項目 ----
        if news:
            ws.cell(row=r, column=C_CODE, value='四').font = normal
            ws.cell(row=r, column=C_DESC, value='新增項目').font = normal
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
            r += 1

            new_data_rows = []

            def emit_new_line(code, desc, unit, qty, unit_price):
                nonlocal r
                ws.cell(row=r, column=C_CODE, value=code).font = normal
                ws.cell(row=r, column=C_DESC, value=desc).font = normal
                ws.cell(row=r, column=C_UNIT, value=unit).font = normal
                oq = ws.cell(row=r, column=C_OQTY, value=0); qty_fmt(oq); oq.font = normal
                nqc = ws.cell(row=r, column=C_NQTY, value=qty); qty_fmt(nqc); nqc.font = normal
                dq = ws.cell(row=r, column=C_DQTY, value=f'=E{r}-D{r}'); qty_fmt(dq); dq.font = normal
                pr = ws.cell(row=r, column=C_PRICE, value=unit_price); pr.number_format = '#,##0.00'; pr.font = normal
                hb = ws.cell(row=r, column=C_BEFORE, value=f'=D{r}*G{r}'); money_fmt(hb); hb.font = normal
                ha = ws.cell(row=r, column=C_AFTER, value=f'=H{r}+J{r}-K{r}'); money_fmt(ha); ha.font = normal
                # 新增項目：J = 全數是新增(E*G)，K = 0
                hi = ws.cell(row=r, column=C_INC, value=f'=E{r}*G{r}'); money_fmt(hi); hi.font = normal
                hd = ws.cell(row=r, column=C_DEC, value=0); money_fmt(hd); hd.font = normal
                for col in range(1, 12):
                    ws.cell(row=r, column=col).border = border
                this = r
                r += 1
                return this

            # <1> 原契約項目
            ws.cell(row=r, column=C_CODE, value='<1>').font = normal
            ws.cell(row=r, column=C_DESC, value='原契約項目').font = normal
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
            r += 1
            for it in news:
                new_data_rows.append(
                    emit_new_line(it.code, it.desc, it.unit, it.qty, it.orig_unit_price))

            # <2> 新增單價部份（一律顯示，即使目前為 0）
            ws.cell(row=r, column=C_CODE, value='<2>').font = normal
            ws.cell(row=r, column=C_DESC, value='新增單價部份').font = normal
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
            r += 1
            for it in news:
                new_data_rows.append(
                    emit_new_line(it.code, it.desc, it.unit, it.qty, it.new_unit_price))

            st = emit_subtotal(new_data_rows)
            section_after_rows.append(st); section_inc_rows.append(st); section_dec_rows.append(st)

        # ---- 合計 ----
        if is_level_group:
            # 本身是層級：不輸出獨立合計，把各段小計登記到此層級，
            # 之後與其子標題群組一起在層級小計合併。
            level_stack[-1]['rows'].extend(section_after_rows)
        else:
            ws.cell(row=r, column=C_CODE, value=g.code).font = bold
            ws.cell(row=r, column=C_DESC, value=f'{g.desc}合計').font = bold
            hb = ws.cell(row=r, column=C_BEFORE, value=group_orig_total); money_fmt(hb); hb.font = bold
            ha = ws.cell(row=r, column=C_AFTER,
                         value='=' + '+'.join(f'I{i}' for i in section_after_rows)); money_fmt(ha); ha.font = bold
            hi = ws.cell(row=r, column=C_INC,
                         value='=' + '+'.join(f'J{i}' for i in section_inc_rows)); money_fmt(hi); hi.font = bold
            hd = ws.cell(row=r, column=C_DEC,
                         value='=' + '+'.join(f'K{i}' for i in section_dec_rows)); money_fmt(hd); hd.font = bold
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = border
                ws.cell(row=r, column=col).fill = grp_fill
            group_total_row = r
            r += 1

            # 把此群組合計登記到最深的開啟層級，供父層小計加總
            if level_stack:
                level_stack[-1]['rows'].append(group_total_row)
            else:
                top_level_rows.append(group_total_row)

    # ---- 收掉所有剩餘層級的小計 ----
    close_levels_to(0)

    # ---- 安全網：把 group_code 對不上任何群組的新增項目補在最後，避免遺漏 ----
    orphans = [it for it in model.new_items if id(it) not in rendered_new_items]
    if orphans:
        ws.cell(row=r, column=C_CODE, value='＊').font = bold
        ws.cell(row=r, column=C_DESC, value='其他新增項目（未歸入既有分組）').font = bold
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = border
            ws.cell(row=r, column=col).fill = grp_fill
        r += 1
        for it in orphans:
            for code, desc, unit, qty, price in (
                    (it.code, it.desc, it.unit, it.qty, it.orig_unit_price),
                    (it.code, it.desc, it.unit, it.qty, it.new_unit_price)):
                ws.cell(row=r, column=C_CODE, value=code).font = normal
                ws.cell(row=r, column=C_DESC, value=desc).font = normal
                ws.cell(row=r, column=C_UNIT, value=unit).font = normal
                oq = ws.cell(row=r, column=C_OQTY, value=0); qty_fmt(oq); oq.font = normal
                nqc = ws.cell(row=r, column=C_NQTY, value=qty); qty_fmt(nqc); nqc.font = normal
                dq = ws.cell(row=r, column=C_DQTY, value=f'=E{r}-D{r}'); qty_fmt(dq); dq.font = normal
                pr = ws.cell(row=r, column=C_PRICE, value=price); pr.number_format = '#,##0.00'; pr.font = normal
                hb = ws.cell(row=r, column=C_BEFORE, value=f'=D{r}*G{r}'); money_fmt(hb); hb.font = normal
                ha = ws.cell(row=r, column=C_AFTER, value=f'=H{r}+J{r}-K{r}'); money_fmt(ha); ha.font = normal
                hi = ws.cell(row=r, column=C_INC, value=f'=E{r}*G{r}'); money_fmt(hi); hi.font = normal
                hd = ws.cell(row=r, column=C_DEC, value=0); money_fmt(hd); hd.font = normal
                for col in range(1, 12):
                    ws.cell(row=r, column=col).border = border
                r += 1

    widths = [16, 44, 6, 11, 11, 10, 10, 16, 16, 14, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # 依標題長度動態放寬 H/I 欄（設上限 34 避免過寬；更長的字交給自動換行處理）
    for col, label in (('H', before_label), ('I', after_label)):
        ws.column_dimensions[col].width = max(16, min(34, int(len(str(label)) * 2.2)))
    ws.freeze_panes = 'A3'

    # ---- 附加分析表 ----
    _add_analysis_sheets(wb, model)
    _add_upa_sheet(wb, model)

    # ---- 全表文字自動換行（保留原水平對齊；垂直預設置中）----
    for _w in wb.worksheets:
        for _row in _w.iter_rows():
            for _cell in _row:
                if _cell.value is None or _cell.value == '':
                    continue
                _al = _cell.alignment
                _cell.alignment = Alignment(
                    horizontal=_al.horizontal,
                    vertical=_al.vertical or 'center',
                    wrap_text=True,
                )

    wb.save(out_path)
    return out_path


# ----------------------------------------------------------------------------
# 分析表：原契約數量增加 / 減少 / 新增項目（比照契約變更分析表格式）
# ----------------------------------------------------------------------------
def _collect_ancestor_chain(model, code):
    """回傳某群組（依 code）由外到內的祖先 (code, desc) 清單。"""
    g = model.group_by_code.get(code)
    return g.ancestors if g else []


def _add_analysis_sheets(wb, model, proj_name='', proj_no=''):
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(name='Arial', bold=True)
    normal = Font(name='Arial')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    title_fill = PatternFill('solid', fgColor='D9E1F2')
    anc_fill = PatternFill('solid', fgColor='F2F2F2')

    # 蒐集三類資料，並保留群組順序與其祖先鏈
    inc_rows, dec_rows, new_rows = [], [], []
    for g in model.groups:
        for lf in g.leaves:
            if lf.code in model.changes:
                nq = model.changes[lf.code]
                if nq > lf.orig_qty:
                    inc_rows.append((g, lf, nq))
                elif nq < lf.orig_qty:
                    dec_rows.append((g, lf, nq))
        for it in model.new_items:
            if it.group_code == g.code:
                new_rows.append((g, it))

    def emit_hierarchy_sheet(ws, headers, records, kind):
        # 標題區
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value='台灣自來水股份有限公司').alignment = center
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value=headers['title']).alignment = center
        ws.cell(row=2, column=1).font = bold
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        ws.cell(row=3, column=1,
                value=f'工程名稱：{proj_name}　工程編號：{proj_no}').alignment = left
        # 欄位標題（第4列）
        for ci, h in enumerate(headers['cols'], start=1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = bold; c.alignment = center; c.border = border; c.fill = title_fill

        r = 5
        printed = set()
        last_anc_key = None
        for rec in records:
            g = rec[0]
            # 印祖先鏈（每條分支變化時才印，避免重複）
            chain = g.ancestors + [(g.code, g.desc)]
            # 只印還沒印過的祖先層
            for code, desc in chain[:-1]:
                if code in printed:
                    continue
                printed.add(code)
                ws.cell(row=r, column=1, value=code).font = bold
                ws.cell(row=r, column=2, value=desc).font = bold
                for ci in range(1, len(headers['cols']) + 1):
                    ws.cell(row=r, column=ci).border = border
                    ws.cell(row=r, column=ci).fill = anc_fill
                r += 1
            # 群組本身當作一個標題（若尚未印）
            gcode = g.code
            if gcode not in printed:
                printed.add(gcode)
                ws.cell(row=r, column=1, value=gcode).font = bold
                ws.cell(row=r, column=2, value=g.desc).font = bold
                for ci in range(1, len(headers['cols']) + 1):
                    ws.cell(row=r, column=ci).border = border
                    ws.cell(row=r, column=ci).fill = anc_fill
                r += 1
            # 明細列
            if kind in ('inc', 'dec'):
                lf, nq = rec[1], rec[2]
                vals = [lf.code, lf.desc, lf.unit,
                        lf.orig_qty, nq, model.reasons.get(lf.code, '')]
            else:  # new
                it = rec[1]
                unit_price = round(it.orig_unit_price + it.new_unit_price, 2)
                vals = [it.code, it.desc, it.unit, unit_price, it.qty, getattr(it, 'reason', '')]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.font = normal; c.border = border
                c.alignment = left if ci == 2 else center
            r += 1

        widths = headers['widths']
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A5'

    ws = wb.create_sheet('原契約數量增加')
    emit_hierarchy_sheet(ws, {
        'title': '工程採購契約變更分析表（一）：原契約數量增加',
        'cols': ['編號', '工項名稱', '單位', '原契約數量', '變更後數量', '數量增加原因分析'],
        'widths': [22, 44, 8, 13, 13, 40],
    }, inc_rows, 'inc')

    ws = wb.create_sheet('原契約數量減少')
    emit_hierarchy_sheet(ws, {
        'title': '工程採購契約變更分析表（二）：原契約數量減少',
        'cols': ['編號', '工項名稱', '單位', '原契約數量', '變更後數量', '數量減少原因分析'],
        'widths': [22, 44, 8, 13, 13, 40],
    }, dec_rows, 'dec')

    ws = wb.create_sheet('新增項目')
    emit_hierarchy_sheet(ws, {
        'title': '工程採購契約變更分析表（三）：新增項目',
        'cols': ['編號', '工項名稱', '單位', '單價', '數量', '新增項目原因分析'],
        'widths': [22, 44, 8, 13, 13, 40],
    }, new_rows, 'new')


def _short_code(code):
    """把完整編碼縮成最後一段顯示（<A><1>[1] -> [1]），分析表用。"""
    import re
    toks = re.findall(r'<[^<>]+>|\[\d+\]|[a-z]|\d+', code)
    return toks[-1] if toks else code


def _leaf_seq(code):
    """明細項次顯示（取最後一段序號）。"""
    return _short_code(code)


# ----------------------------------------------------------------------------
# 單價分析表工作表：把所有新增項目的單價分析整理成一張（可微調）
# ----------------------------------------------------------------------------
def _add_upa_sheet(wb, model):
    news = list(model.new_items)
    if not news:
        return
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(name='Arial', bold=True)
    normal = Font(name='Arial')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    sec_fill = PatternFill('solid', fgColor='FFF2CC')
    title_fill = PatternFill('solid', fgColor='FCE4D6')

    ws = wb.create_sheet('單價分析表')
    cols = ['項次', '工料項目及說明', '單位', '數量', '單價', '複價', '備註']
    widths = [8, 44, 8, 10, 12, 14, 16]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 備註欄下拉選單（物調材料類別）
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type='list',
                        formula1='"' + ','.join(MATERIAL_TYPES) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    red = Font(name='Arial', color='FF0000')
    red_bold = Font(name='Arial', bold=True, color='FF0000')

    r = 1
    for it in news:
        # 項目標題
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c = ws.cell(row=r, column=1,
                    value=f'{it.code}　{it.desc}　（單位：{it.unit}　數量：{it.qty}）')
        c.font = bold; c.fill = title_fill; c.alignment = left
        for ci in range(1, 8):
            ws.cell(row=r, column=ci).border = border
        r += 1
        # 欄位標題
        for ci, h in enumerate(cols, start=1):
            cc = ws.cell(row=r, column=ci, value=h)
            cc.font = bold; cc.alignment = center; cc.border = border; cc.fill = hdr_fill
        r += 1

        mat_orig = {}   # 原契約部份: 材料 -> [複價列...]
        mat_new = {}    # 新增部份:   材料 -> [複價列...]

        def emit_part(title, rows, mdict):
            nonlocal r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            cc = ws.cell(row=r, column=1, value=title)
            cc.font = bold; cc.fill = sec_fill; cc.alignment = left
            for ci in range(1, 8):
                ws.cell(row=r, column=ci).border = border
            r += 1
            first = r
            item_num = 1
            for row in rows:
                ws.cell(row=r, column=1, value=item_num).font = normal
                ws.cell(row=r, column=2, value=row.desc).font = normal
                ws.cell(row=r, column=3, value=row.unit).font = normal
                ws.cell(row=r, column=4, value=row.qty).font = normal
                pcell = ws.cell(row=r, column=5, value=row.price); pcell.font = normal
                pcell.number_format = '#,##0.00'
                tcell = ws.cell(row=r, column=6, value=f'=D{r}*E{r}')
                tcell.number_format = '#,##0.00'; tcell.font = normal
                rk = ws.cell(row=r, column=7, value=(row.remark or None))
                rk.font = red if row.remark else normal
                dv.add(rk)   # 備註欄掛下拉
                if row.remark:
                    mdict.setdefault(row.remark, []).append(r)
                for ci in range(1, 8):
                    ws.cell(row=r, column=ci).border = border
                r += 1
                item_num += 1
            last = r - 1
            # 小計
            ws.cell(row=r, column=2, value='小計').font = bold
            sc = ws.cell(row=r, column=6,
                         value=(f'=SUM(F{first}:F{last})' if rows else 0))
            sc.number_format = '#,##0.00'; sc.font = bold
            for ci in range(1, 8):
                ws.cell(row=r, column=ci).border = border
            sub_row = r
            r += 1
            # 物調權重（緊接在該區小計下方；分子=材料複價、分母=該區小計）
            for mat in MATERIAL_TYPES:
                if mat not in mdict:
                    continue
                num = '+'.join(f'F{i}' for i in mdict[mat])
                # 標籤用活公式顯示完整算式：物調權重：鋼筋（170,765/222,298）＝
                lbl = ws.cell(row=r, column=2,
                              value=(f'="物調權重："&"{mat}"&"（"&TEXT({num},"#,##0.00")'
                                     f'&"/"&TEXT(F{sub_row},"#,##0.00")&"）＝"'))
                lbl.font = red_bold
                wc = ws.cell(row=r, column=6,
                             value=f'=IF(F{sub_row}=0,0,({num})/F{sub_row})')
                wc.number_format = '0.0000%'; wc.font = red_bold
                for ci in range(1, 8):
                    ws.cell(row=r, column=ci).border = border
                r += 1
            return sub_row

        o_sub = emit_part('一、原契約單價部份', it.upa_orig, mat_orig)
        n_sub = emit_part('二、新增單價部份', it.upa_new, mat_new)
        # 合計單價
        ws.cell(row=r, column=2, value='每單位單價合計').font = bold
        tc = ws.cell(row=r, column=6, value=f'=F{o_sub}+F{n_sub}')
        tc.number_format = '#,##0.00'; tc.font = bold
        for ci in range(1, 8):
            ws.cell(row=r, column=ci).border = border
            ws.cell(row=r, column=ci).fill = title_fill
        r += 2  # 空一列
